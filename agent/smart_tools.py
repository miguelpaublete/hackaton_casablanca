"""
smart_tools.py — Herramientas inteligentes del KDD PMO Copilot.

1. Smart Agenda: Genera orden del día sugerida antes de reuniones.
2. Auto-Setup: Genera planificación inicial de proyecto desde un prompt.
3. Onboarding Kit: Genera kit de bienvenida para nuevos integrantes.
"""

import json
import re
from dataclasses import dataclass
from datetime import date
from pathlib import Path

import config
from acta_loader import read_acta
from extractor import call_github_models


# ─────────────────────────────────────────────────────────────
# PRESUPUESTO DE CONTEXTO
# ────────��────────────────────────────────────────────────────
# Sistema de activación por tiers (basado en unified-taxonomy.md):
#   Tier 1 — contenido completo:  proyecto propio (actas + specs)
#   Tier 2 — resumen compacto:    otros proyectos (frontmatter + Intent)
#   Tier 3 — solo referencia:     lo que no cabe (IDs listados, no inyectados)
#
# Esto garantiza que el contexto no explote cuando el repositorio
# global crezca a 50+ proyectos con cientos de specs.

BUDGET = {
    # Chars máximos por item
    "acta_per_file":        3_000,
    "spec_own_project":     2_000,
    "spec_other_project":     500,   # Tier 2: solo resumen
    "historical_plan":      3_000,
    # Totales máximos por función
    "smart_agenda_total":  20_000,
    "setup_global_specs":  10_000,
    "setup_historical":    15_000,
    "onboarding_project":  20_000,
    "onboarding_global":    8_000,
    # Nº máximo de items
    "max_actas":                5,
    "max_own_specs":           15,
    "max_other_specs":         20,
    "max_historical_plans":     5,
}

# ─────────────────────────────────────────────────────────────
# PATHS — Dual source:
#   LOCAL_SPECS_DIR  → specs generadas por el PMO desde actas (../specs)
#   GLOBAL_SPECS_DIR → specs del KDD global del equipo (cib-risk-knowledge/specs/)
#   El agente combina AMBAS fuentes para dar contexto más rico.
# ─────────────────────────────────────────────────────────────

_PMO_ROOT = config.PROJECT_ROOT.parent

# Specs locales del PMO (generadas por el agente desde actas)
LOCAL_SPECS_DIR = _PMO_ROOT / "specs"

# Specs globales del equipo KDD (auto-detectar si estamos dentro de cib-risk-knowledge/projects/PMO/)
_candidate_global = _PMO_ROOT.parent.parent / "specs"
if (_candidate_global.exists()
    and _candidate_global.resolve() != LOCAL_SPECS_DIR.resolve()
    and any(_candidate_global.iterdir())):
    GLOBAL_SPECS_DIR: Path | None = _candidate_global
else:
    GLOBAL_SPECS_DIR = None

# Compatibilidad: SPECS_DIR = local (donde se guardan las specs del PMO)
SPECS_DIR = LOCAL_SPECS_DIR

ACTAS_DIR            = _PMO_ROOT / "actas"
SMART_AGENDA_DIR     = _PMO_ROOT / "Smart agenda"
SETUP_DIR            = _PMO_ROOT / "Set-up proyecto"
HISTORICAL_PLANS_DIR = SETUP_DIR / "_historical"


# ─────────────────────────────────────────────────────────────
# UTILIDADES DE RELEVANCIA Y CONTEXTO
# ─────────────────────────────────────────────────────────────

def _slugify(text: str) -> str:
    text = re.sub(r"[^\w\s-]", "", text, flags=re.UNICODE).strip().lower()
    return re.sub(r"[-\s]+", "-", text)[:60] or "sin-nombre"


def _extract_keywords(text: str) -> list[str]:
    """Extrae palabras clave significativas de un texto para filtrar por relevancia."""
    stopwords = {
        "de", "la", "el", "en", "y", "a", "los", "las", "por", "para",
        "con", "que", "un", "una", "es", "se", "del", "al", "le", "su",
        "the", "an", "of", "in", "and", "to", "for", "with", "is", "are",
        "was", "were", "be", "been", "have", "has", "had",
    }
    words = re.findall(r"\b[a-záéíóúñA-ZÁÉÍÓÚÑ]{4,}\b", text)
    return list({w.lower() for w in words if w.lower() not in stopwords})[:30]


def _score_relevance(content: str, keywords: list[str]) -> int:
    """Cuenta cuántas keywords aparecen en el contenido (case-insensitive)."""
    content_lower = content.lower()
    return sum(1 for kw in keywords if kw.lower() in content_lower)


def _extract_spec_summary(content: str, max_chars: int = 500) -> str:
    """
    Tier 2: resumen compacto de una spec.
    Conserva solo campos clave del frontmatter + sección Intent (máx 5 líneas).
    """
    lines = content.splitlines()
    summary_lines = []
    in_frontmatter = False
    past_frontmatter = False
    intent_lines = 0
    in_intent = False

    for line in lines:
        if line.strip() == "---" and not past_frontmatter:
            in_frontmatter = not in_frontmatter
            if not in_frontmatter:
                past_frontmatter = True
            summary_lines.append(line)
            continue

        if in_frontmatter:
            key = line.split(":")[0].strip()
            if key in ("id", "type", "layer", "domain", "subdomain",
                       "status", "confidence", "owner", "project", "tags"):
                summary_lines.append(line)
            continue

        if past_frontmatter:
            if line.startswith("## Intent") or line.startswith("## Objetivo"):
                in_intent = True
                summary_lines.append(line)
                continue
            if in_intent:
                if line.startswith("## ") and intent_lines > 0:
                    break
                summary_lines.append(line)
                intent_lines += 1
                if intent_lines >= 5:
                    break

    return "\n".join(summary_lines)[:max_chars]


def _gather_project_context(
    project: str,
    max_actas: int | None = None,
    relevance_keywords: list[str] | None = None,
    budget_chars: int | None = None,
) -> str:
    """
    Tier 1: contexto completo del proyecto (actas + specs propias).
    Las specs se ordenan por relevancia si se proporcionan keywords.
    Respeta el presupuesto total de chars.
    """
    max_actas = max_actas or BUDGET["max_actas"]
    budget = budget_chars or BUDGET["smart_agenda_total"]
    used_chars = 0
    chunks = []

    # — Últimas actas del proyecto —
    actas_path = ACTAS_DIR / project
    if actas_path.exists():
        acta_files = sorted(
            [f for f in actas_path.iterdir()
             if f.suffix in (".txt", ".md", ".pdf", ".docx")
             and not f.name.endswith(".converted.txt")
             and f.name != "README.md"],
            key=lambda f: f.stat().st_mtime,
            reverse=True,
        )[:max_actas]
        for af in acta_files:
            if used_chars >= budget:
                break
            text = read_acta(af)[:BUDGET["acta_per_file"]]
            chunk = f"### Acta: {af.name}\n{text}\n"
            chunks.append(chunk)
            used_chars += len(chunk)

    # — Specs del proyecto (Tier 1: contenido completo) —
    specs_path = SPECS_DIR / project
    if specs_path.exists():
        spec_files = list(specs_path.rglob("*.md"))
        if relevance_keywords:
            spec_files.sort(
                key=lambda f: _score_relevance(
                    f.read_text(encoding="utf-8", errors="ignore"),
                    relevance_keywords,
                ),
                reverse=True,
            )
        for sf in spec_files[:BUDGET["max_own_specs"]]:
            if used_chars >= budget:
                break
            content = sf.read_text(encoding="utf-8", errors="ignore")[:BUDGET["spec_own_project"]]
            chunk = f"### Spec: {sf.name}\n{content}\n"
            chunks.append(chunk)
            used_chars += len(chunk)

    print(f"  [Tier 1 — {project}] {len(chunks)} items, ~{used_chars:,} chars")
    return "\n---\n".join(chunks) if chunks else "(Sin contexto previo disponible para este proyecto)"


def _gather_global_specs_tiered(
    relevance_keywords: list[str],
    exclude_project: str = "",
    budget_chars: int = 10_000,
) -> str:
    """
    Tier 2: resúmenes compactos de specs de OTROS proyectos,
    ordenados por relevancia y limitados por presupuesto.
    Busca en AMBAS fuentes: LOCAL_SPECS_DIR + GLOBAL_SPECS_DIR.
    Lo que no cabe va al Tier 3 (solo listado de IDs).
    """
    used_chars = 0
    chunks = []
    tier3_ids = []

    # Construir lista de directorios de specs a escanear
    specs_sources: list[tuple[str, Path]] = []
    if LOCAL_SPECS_DIR.exists():
        specs_sources.append(("PMO", LOCAL_SPECS_DIR))
    if GLOBAL_SPECS_DIR and GLOBAL_SPECS_DIR.exists():
        specs_sources.append(("KDD", GLOBAL_SPECS_DIR))

    if not specs_sources:
        return "(Sin specs disponibles — ni locales ni globales)"

    # Recopilar y puntuar specs de AMBAS fuentes
    all_foreign: list[tuple[int, str, Path, str]] = []
    for source_label, source_dir in specs_sources:
        for item in sorted(source_dir.iterdir()):
            if not item.is_dir() or item.name == exclude_project:
                continue
            # Saltar carpetas ocultas y README
            if item.name.startswith(".") or item.name.startswith("_"):
                continue
            for sf in item.rglob("*.md"):
                if sf.name == "README.md":
                    continue
                content = sf.read_text(encoding="utf-8", errors="ignore")
                score = _score_relevance(content, relevance_keywords)
                # Etiquetar según fuente para claridad en el output
                proj_label = f"{source_label}/{item.name}"
                all_foreign.append((score, proj_label, sf, content))

    all_foreign.sort(key=lambda x: x[0], reverse=True)

    for score, proj_name, sf, content in all_foreign[:BUDGET["max_other_specs"]]:
        summary = _extract_spec_summary(content, BUDGET["spec_other_project"])
        chunk = f"### [{proj_name}] {sf.name} (relevancia: {score})\n{summary}\n"

        if used_chars + len(chunk) > budget_chars:
            id_match = re.search(r"^id:\s*(.+)$", content, re.MULTILINE)
            spec_id = id_match.group(1).strip() if id_match else sf.stem
            tier3_ids.append(f"{spec_id} [{proj_name}]")
            continue

        chunks.append(chunk)
        used_chars += len(chunk)

    result = "\n---\n".join(chunks) if chunks else ""
    if tier3_ids:
        result += (
            "\n\n> **Specs adicionales (Tier 3 — disponibles pero no incluidas "
            "por presupuesto):** " + ", ".join(tier3_ids)
        )

    sources_str = " + ".join(sl for sl, _ in specs_sources)
    print(f"  [Tier 2 — {sources_str}] {len(chunks)} specs (~{used_chars:,} chars), "
          f"{len(tier3_ids)} en Tier 3")
    return result or "(Sin specs globales relevantes)"


def _gather_historical_plans(
    relevance_keywords: list[str] | None = None,
    budget_chars: int = 15_000,
) -> str:
    """
    Planificaciones históricas de referencia, ordenadas por relevancia
    y limitadas por presupuesto.
    """
    if not HISTORICAL_PLANS_DIR.exists():
        return "(Sin planes históricos cargados)"

    used_chars = 0
    chunks = []
    plan_files = (
        list(HISTORICAL_PLANS_DIR.glob("*.md")) +
        list(HISTORICAL_PLANS_DIR.glob("*.txt"))
    )

    if relevance_keywords:
        plan_files.sort(
            key=lambda f: _score_relevance(
                f.read_text(encoding="utf-8", errors="ignore"),
                relevance_keywords,
            ),
            reverse=True,
        )

    for f in plan_files[:BUDGET["max_historical_plans"]]:
        if used_chars >= budget_chars:
            break
        content = f.read_text(encoding="utf-8", errors="ignore")[:BUDGET["historical_plan"]]
        chunk = f"### Plan histórico: {f.name}\n{content}\n"
        chunks.append(chunk)
        used_chars += len(chunk)

    print(f"  [Histórico] {len(chunks)} planes, ~{used_chars:,} chars")
    return "\n---\n".join(chunks) if chunks else "(Sin planes históricos cargados)"


# ─────────────────────────────────────────────────────────────
# 1. SMART AGENDA
# ─────────────────────────────────────────────────────────────

@dataclass
class AgendaResult:
    project: str
    agenda_md: str
    key_points: list[str]
    pending_decisions: list[str]
    risks: list[str]
    saved_path: Path | None = None


def generate_smart_agenda(
    project: str,
    meeting_type: str = "seguimiento semanal",
    additional_context: str = "",
) -> AgendaResult:
    """
    Genera una orden del día inteligente basada en el contexto del proyecto.
    """
    keywords = _extract_keywords(f"{project} {meeting_type} {additional_context}")
    project_context = _gather_project_context(
        project,
        max_actas=3,
        relevance_keywords=keywords,
        budget_chars=BUDGET["smart_agenda_total"],
    )

    prompt = f"""You are an expert PMO assistant. Generate a smart meeting agenda in Spanish.

## CONTEXT
- Project: {project}
- Meeting type: {meeting_type}
- Today: {date.today().isoformat()}
- Additional notes from PMO: {additional_context or 'None'}

## PROJECT HISTORY (last meetings and specs):
{project_context}

## TASK
Analyze the project history and generate:
1. A suggested agenda (ordered list of discussion points with time estimates)
2. Key points that MUST be discussed (blockers, pending decisions, deadlines)
3. Pending decisions that need resolution
4. Identified risks

## OUTPUT FORMAT (JSON only, no markdown fences):
{{
  "agenda_md": "# Orden del Día - {{project}}\\n\\n## 1. Punto...\\n...(full markdown agenda)",
  "key_points": ["point1", "point2"],
  "pending_decisions": ["decision1", "decision2"],
  "risks": ["risk1", "risk2"]
}}
"""
    raw = call_github_models(prompt)
    cleaned = raw.strip()
    if cleaned.startswith("```"):
        cleaned = re.sub(r"^```(?:json)?\s*\n?", "", cleaned)
        cleaned = re.sub(r"\n?```\s*$", "", cleaned)

    data = json.loads(cleaned)
    result = AgendaResult(
        project=project,
        agenda_md=data.get("agenda_md", ""),
        key_points=data.get("key_points", []),
        pending_decisions=data.get("pending_decisions", []),
        risks=data.get("risks", []),
    )

    # Guardar como .txt en Smart agenda/{project}/
    project_dir = SMART_AGENDA_DIR / project
    project_dir.mkdir(parents=True, exist_ok=True)
    today_str = date.today().isoformat()
    out_path = project_dir / f"smart_agenda_reunion_{today_str}.txt"

    body = []
    body.append(f"SMART AGENDA — Reunión del día {today_str}")
    body.append(f"Proyecto: {project}")
    body.append(f"Tipo de reunión: {meeting_type}")
    body.append("=" * 70)
    body.append("")
    body.append(result.agenda_md)
    body.append("")
    body.append("-" * 70)
    body.append("PUNTOS CLAVE:")
    for p in result.key_points:
        body.append(f"  • {p}")
    body.append("")
    body.append("DECISIONES PENDIENTES:")
    for d in result.pending_decisions:
        body.append(f"  • {d}")
    body.append("")
    body.append("RIESGOS IDENTIFICADOS:")
    for r in result.risks:
        body.append(f"  • {r}")

    out_path.write_text("\n".join(body), encoding="utf-8")
    result.saved_path = out_path
    return result


# ─────────────────────────────────────────────────────────────
# 2. AUTO-SETUP DE PROYECTO
# ─────────────────────────────────────────────────────────────

@dataclass
class ProjectSetupResult:
    project_name: str
    wbs_md: str
    deliverables: list[str]
    milestones: list[dict]
    suggested_structure: str
    wbs_tasks: list[dict]  # [{phase, task, owner, week_start, week_end, duration}]
    saved_path: Path | None = None


def generate_project_setup(
    project_description: str,
    project_name: str = "",
    duration_weeks: int = 12,
) -> ProjectSetupResult:
    """
    Genera la planificación inicial y estructura documental de un proyecto.
    Consulta specs históricas + planificaciones históricas como referencia.
    Exporta a Excel en `Set-up proyecto/`.
    """
    keywords = _extract_keywords(f"{project_name} {project_description}")
    historical_plans = _gather_historical_plans(
        relevance_keywords=keywords,
        budget_chars=BUDGET["setup_historical"],
    )
    historical_specs = _gather_global_specs_tiered(
        relevance_keywords=keywords,
        budget_chars=BUDGET["setup_global_specs"],
    )

    prompt = f"""You are a senior PMO in a banking consultancy (BBVA). Generate a complete project setup in Spanish.

## INPUT
- Project name: {project_name or 'To be defined from description'}
- Description: {project_description}
- Estimated duration: {duration_weeks} weeks
- Today: {date.today().isoformat()}

## HISTORICAL PROJECT PLANS (primary reference — use their structure and patterns):
{historical_plans}

## RELEVANT SPECS FROM OTHER PROJECTS (secondary context, Tier 2 summaries):
{historical_specs}

## TASK
Generate:
1. A Work Breakdown Structure (WBS) in Markdown with phases, tasks, and subtasks
2. List of required deliverables
3. Critical milestones with estimated dates
4. Suggested folder/document structure for the project
5. A FLAT list of WBS tasks with owner, week_start, week_end and duration_weeks (for Excel export)

Use patterns from the historical plans when applicable.

## OUTPUT FORMAT (JSON only):
{{
  "project_name": "Final project name",
  "wbs_md": "# WBS\\n\\n## Fase 1...\\n...(complete WBS in markdown)",
  "deliverables": ["deliverable1", "deliverable2"],
  "milestones": [
    {{"name": "Kickoff", "week": 1, "description": "..."}},
    {{"name": "MVP", "week": 6, "description": "..."}}
  ],
  "suggested_structure": "```\\nproject/\\n  specs/\\n  docs/\\n  ...\\n```",
  "wbs_tasks": [
    {{"phase": "Fase 1 - Análisis", "task": "Recogida de requisitos", "owner": "PM", "week_start": 1, "week_end": 2, "duration_weeks": 2}},
    {{"phase": "Fase 1 - Análisis", "task": "Aprobación stakeholders", "owner": "PMO", "week_start": 2, "week_end": 3, "duration_weeks": 1}}
  ]
}}
"""
    raw = call_github_models(prompt)
    cleaned = raw.strip()
    if cleaned.startswith("```"):
        cleaned = re.sub(r"^```(?:json)?\s*\n?", "", cleaned)
        cleaned = re.sub(r"\n?```\s*$", "", cleaned)

    data = json.loads(cleaned)
    result = ProjectSetupResult(
        project_name=data.get("project_name", project_name),
        wbs_md=data.get("wbs_md", ""),
        deliverables=data.get("deliverables", []),
        milestones=data.get("milestones", []),
        suggested_structure=data.get("suggested_structure", ""),
        wbs_tasks=data.get("wbs_tasks", []),
    )

    # Generar Excel en Set-up proyecto/
    result.saved_path = _export_setup_to_excel(result, duration_weeks)
    return result


def _export_setup_to_excel(result: ProjectSetupResult, duration_weeks: int) -> Path:
    """
    Genera un Excel profesional con:
    - Resumen del proyecto
    - WBS con seguimiento (estado, % avance, dependencias)
    - Gantt Chart interactivo (barras automáticas por semana)
    - Entregables, Hitos, Estructura
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
        from openpyxl.worksheet.datavalidation import DataValidation
        from openpyxl.formatting.rule import CellIsRule, FormulaRule
        from openpyxl.utils import get_column_letter
    except ImportError:
        # Fallback: guardar como markdown si no hay openpyxl
        SETUP_DIR.mkdir(parents=True, exist_ok=True)
        slug = _slugify(result.project_name)
        out = SETUP_DIR / f"setup_{slug}_{date.today().isoformat()}.md"
        out.write_text(
            f"# {result.project_name}\n\n## WBS\n{result.wbs_md}\n\n"
            f"## Entregables\n" + "\n".join(f"- {d}" for d in result.deliverables) +
            f"\n\n## Hitos\n" + "\n".join(f"- Semana {m.get('week')}: {m.get('name')} — {m.get('description','')}" for m in result.milestones) +
            f"\n\n## Estructura\n{result.suggested_structure}",
            encoding="utf-8",
        )
        return out

    SETUP_DIR.mkdir(parents=True, exist_ok=True)
    slug = _slugify(result.project_name)
    out_path = SETUP_DIR / f"setup_{slug}_{date.today().isoformat()}.xlsx"

    wb = Workbook()

    # ── Estilos comunes ──
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    subheader_fill = PatternFill("solid", fgColor="2E75B6")
    thin_border = Border(
        left=Side("thin", color="D9D9D9"),
        right=Side("thin", color="D9D9D9"),
        top=Side("thin", color="D9D9D9"),
        bottom=Side("thin", color="D9D9D9"),
    )

    # Colores para estados
    STATUS_COLORS = {
        "Pendiente": "BDD7EE",
        "En Plan": "D6E4F0",
        "En Progreso": "C6EFCE",
        "Completado": "A9D18E",
        "Bloqueado": "FFC7CE",
        "Retrasado": "F4B183",
        "En Revisión": "FFE699",
        "Cancelado": "D9D9D9",
    }
    GANTT_FILL = PatternFill("solid", fgColor="2E75B6")
    GANTT_MILESTONE = PatternFill("solid", fgColor="FF6B35")
    GANTT_EMPTY = PatternFill("solid", fgColor="F2F2F2")

    # ════════════════════════════════════════════════════════════
    # HOJA 1: RESUMEN
    # ════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Resumen"
    ws["A1"] = "PLANIFICACIÓN DE PROYECTO"
    ws["A1"].font = Font(bold=True, size=16, color="1F4E78")
    ws.merge_cells("A1:D1")

    info = [
        ("Proyecto:", result.project_name),
        ("Duración:", f"{duration_weeks} semanas"),
        ("Fecha generación:", date.today().isoformat()),
        ("Tareas:", str(len(result.wbs_tasks))),
        ("Entregables:", str(len(result.deliverables))),
        ("Hitos críticos:", str(len(result.milestones))),
    ]
    for i, (label, value) in enumerate(info, 3):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws.cell(row=i, column=2, value=value)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 60

    # Leyenda de estados
    ws.cell(row=11, column=1, value="LEYENDA DE ESTADOS").font = Font(bold=True, size=12)
    for i, (status, color) in enumerate(STATUS_COLORS.items(), 12):
        ws.cell(row=i, column=1, value=status)
        ws.cell(row=i, column=1).fill = PatternFill("solid", fgColor=color)

    # ════════════════════════════════════════════════════════════
    # HOJA 2: SEGUIMIENTO (WBS + estado + % + dependencias)
    # ════════════════════════════════════════════════════════════
    ws_track = wb.create_sheet("Seguimiento")
    track_headers = [
        "ID", "Fase", "Tarea", "Responsable",
        "Estado", "% Avance", "Dependencias",
        "Semana inicio", "Semana fin", "Duración (sem)",
        "Notas / Comentarios"
    ]
    for col, h in enumerate(track_headers, 1):
        c = ws_track.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = thin_border

    # Data validation: desplegable de estado
    status_options = ",".join(STATUS_COLORS.keys())
    dv_status = DataValidation(
        type="list",
        formula1=f'"{status_options}"',
        allow_blank=True,
    )
    dv_status.error = "Selecciona un estado válido"
    dv_status.errorTitle = "Estado no válido"
    dv_status.prompt = "Selecciona el estado de la tarea"
    dv_status.promptTitle = "Estado"
    ws_track.add_data_validation(dv_status)

    # Data validation: % avance (0-100)
    dv_pct = DataValidation(
        type="whole",
        operator="between",
        formula1="0",
        formula2="100",
        allow_blank=True,
    )
    dv_pct.error = "Introduce un valor entre 0 y 100"
    dv_pct.errorTitle = "Porcentaje no válido"
    ws_track.add_data_validation(dv_pct)

    num_tasks = len(result.wbs_tasks)
    for i, t in enumerate(result.wbs_tasks, 2):
        task_id = f"T-{i-1:03d}"
        ws_track.cell(row=i, column=1, value=task_id)
        ws_track.cell(row=i, column=2, value=t.get("phase", ""))
        ws_track.cell(row=i, column=3, value=t.get("task", ""))
        ws_track.cell(row=i, column=4, value=t.get("owner", ""))
        ws_track.cell(row=i, column=5, value="Pendiente")  # Estado inicial
        ws_track.cell(row=i, column=6, value=0)  # % Avance inicial
        ws_track.cell(row=i, column=7, value="")  # Dependencias (el PMO las rellena: T-001, T-002)
        ws_track.cell(row=i, column=8, value=t.get("week_start", ""))
        ws_track.cell(row=i, column=9, value=t.get("week_end", ""))
        ws_track.cell(row=i, column=10, value=t.get("duration_weeks", ""))
        ws_track.cell(row=i, column=11, value="")  # Notas

        # Aplicar bordes
        for col in range(1, 12):
            ws_track.cell(row=i, column=col).border = thin_border
            ws_track.cell(row=i, column=col).alignment = Alignment(vertical="center", wrap_text=True)

    # Aplicar validaciones a rangos
    last_row = num_tasks + 1
    dv_status.add(f"E2:E{last_row}")
    dv_pct.add(f"F2:F{last_row}")

    # Formato condicional por estado
    for status, color in STATUS_COLORS.items():
        ws_track.conditional_formatting.add(
            f"E2:E{last_row}",
            CellIsRule(
                operator="equal",
                formula=[f'"{status}"'],
                fill=PatternFill("solid", fgColor=color),
            ),
        )

    # Formato condicional: % avance con barra de color (verde gradual)
    ws_track.conditional_formatting.add(
        f"F2:F{last_row}",
        CellIsRule(
            operator="greaterThanOrEqual",
            formula=["100"],
            fill=PatternFill("solid", fgColor="A9D18E"),
            font=Font(bold=True, color="375623"),
        ),
    )
    ws_track.conditional_formatting.add(
        f"F2:F{last_row}",
        CellIsRule(
            operator="between",
            formula=["50", "99"],
            fill=PatternFill("solid", fgColor="C6EFCE"),
        ),
    )
    ws_track.conditional_formatting.add(
        f"F2:F{last_row}",
        CellIsRule(
            operator="between",
            formula=["1", "49"],
            fill=PatternFill("solid", fgColor="FFE699"),
        ),
    )

    # Anchos de columna
    col_widths = [8, 26, 42, 16, 14, 12, 18, 14, 12, 14, 30]
    for i, w in enumerate(col_widths, 1):
        ws_track.column_dimensions[get_column_letter(i)].width = w

    # Congelar panel
    ws_track.freeze_panes = "D2"

    # ════════════════════════════════════════════════════════════
    # HOJA 3: GANTT CHART
    # ════════════════════════════════════════════════════════════
    ws_gantt = wb.create_sheet("Gantt Chart")

    # Cabecera izquierda
    gantt_left_headers = ["ID", "Tarea", "Responsable", "Estado", "Inicio", "Fin"]
    gantt_left_cols = len(gantt_left_headers)
    gantt_week_start_col = gantt_left_cols + 1  # Columna donde empiezan las semanas

    for col, h in enumerate(gantt_left_headers, 1):
        c = ws_gantt.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border

    # Cabecera de semanas
    for week in range(1, duration_weeks + 1):
        col = gantt_week_start_col + week - 1
        c = ws_gantt.cell(row=1, column=col, value=f"S{week}")
        c.font = Font(bold=True, color="FFFFFF", size=9)
        c.fill = subheader_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border
        ws_gantt.column_dimensions[get_column_letter(col)].width = 4

    # Anchos columnas izquierda
    for i, w in enumerate([8, 40, 14, 13, 7, 7], 1):
        ws_gantt.column_dimensions[get_column_letter(i)].width = w

    # Filas de tareas con barras de Gantt
    for i, t in enumerate(result.wbs_tasks, 2):
        task_id = f"T-{i-1:03d}"
        ws_gantt.cell(row=i, column=1, value=task_id).alignment = Alignment(horizontal="center")
        ws_gantt.cell(row=i, column=2, value=t.get("task", ""))
        ws_gantt.cell(row=i, column=3, value=t.get("owner", ""))

        # Estado - referencia a la hoja Seguimiento para que se sincronice
        status_formula = f"=Seguimiento!E{i}"
        ws_gantt.cell(row=i, column=4, value=status_formula)
        ws_gantt.cell(row=i, column=4).alignment = Alignment(horizontal="center")

        week_start = t.get("week_start", 1)
        week_end = t.get("week_end", week_start)
        ws_gantt.cell(row=i, column=5, value=week_start).alignment = Alignment(horizontal="center")
        ws_gantt.cell(row=i, column=6, value=week_end).alignment = Alignment(horizontal="center")

        # Bordes izquierda
        for col in range(1, gantt_left_cols + 1):
            ws_gantt.cell(row=i, column=col).border = thin_border

        # Dibujar barras de Gantt
        for week in range(1, duration_weeks + 1):
            col = gantt_week_start_col + week - 1
            cell = ws_gantt.cell(row=i, column=col)
            cell.border = thin_border

            if week_start <= week <= week_end:
                cell.fill = GANTT_FILL
                # Marcar inicio y fin con texto
                if week == week_start and week == week_end:
                    cell.value = "◆"
                    cell.fill = GANTT_MILESTONE
                    cell.font = Font(color="FFFFFF", size=8)
                    cell.alignment = Alignment(horizontal="center")
            else:
                cell.fill = GANTT_EMPTY

    # Agregar hitos como filas especiales
    milestone_start_row = len(result.wbs_tasks) + 3
    ws_gantt.cell(row=milestone_start_row, column=1, value="HITOS").font = Font(bold=True, size=11, color="FF6B35")
    ws_gantt.merge_cells(
        start_row=milestone_start_row, start_column=1,
        end_row=milestone_start_row, end_column=gantt_left_cols,
    )

    for mi, m in enumerate(result.milestones, milestone_start_row + 1):
        ws_gantt.cell(row=mi, column=1, value="⚑")
        ws_gantt.cell(row=mi, column=2, value=m.get("name", ""))
        ws_gantt.cell(row=mi, column=3, value="")
        ws_gantt.cell(row=mi, column=4, value="Hito")
        m_week = m.get("week", 1)
        ws_gantt.cell(row=mi, column=5, value=m_week)
        ws_gantt.cell(row=mi, column=6, value=m_week)

        for col in range(1, gantt_left_cols + 1):
            ws_gantt.cell(row=mi, column=col).border = thin_border
            ws_gantt.cell(row=mi, column=col).font = Font(italic=True, color="FF6B35")

        # Dibujar diamante del hito
        for week in range(1, duration_weeks + 1):
            col = gantt_week_start_col + week - 1
            cell = ws_gantt.cell(row=mi, column=col)
            cell.border = thin_border
            if week == m_week:
                cell.value = "◆"
                cell.fill = GANTT_MILESTONE
                cell.font = Font(color="FFFFFF", bold=True, size=10)
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.fill = GANTT_EMPTY

    # Formato condicional para el estado en Gantt
    gantt_last_row = milestone_start_row + len(result.milestones)
    for status, color in STATUS_COLORS.items():
        ws_gantt.conditional_formatting.add(
            f"D2:D{gantt_last_row}",
            CellIsRule(
                operator="equal",
                formula=[f'"{status}"'],
                fill=PatternFill("solid", fgColor=color),
            ),
        )

    # Congelar panel
    ws_gantt.freeze_panes = f"{get_column_letter(gantt_week_start_col)}2"

    # Altura de filas
    for row in range(1, gantt_last_row + 1):
        ws_gantt.row_dimensions[row].height = 22

    # ════════════════════════════════════════════════════════════
    # HOJA 4: ENTREGABLES
    # ════════════════════════════════════════════════════════════
    ws_d = wb.create_sheet("Entregables")
    del_headers = ["#", "Entregable", "Responsable", "Semana entrega", "Estado"]
    for col, h in enumerate(del_headers, 1):
        c = ws_d.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.border = thin_border

    dv_status_d = DataValidation(type="list", formula1=f'"{status_options}"', allow_blank=True)
    ws_d.add_data_validation(dv_status_d)

    for i, d in enumerate(result.deliverables, 2):
        ws_d.cell(row=i, column=1, value=i - 1)
        ws_d.cell(row=i, column=2, value=d)
        ws_d.cell(row=i, column=3, value="")  # Responsable editable
        ws_d.cell(row=i, column=4, value="")  # Semana editable
        ws_d.cell(row=i, column=5, value="Pendiente")
        for col in range(1, 6):
            ws_d.cell(row=i, column=col).border = thin_border

    dv_status_d.add(f"E2:E{len(result.deliverables) + 1}")
    for col_letter, w in zip("ABCDE", (5, 60, 18, 16, 14)):
        ws_d.column_dimensions[col_letter].width = w

    # ══════════════════════════════════════════════════════��═════
    # HOJA 5: HITOS
    # ════════════════════════════════════════════════════════════
    ws_m = wb.create_sheet("Hitos")
    for col, h in enumerate(["Semana", "Hito", "Descripción", "Estado"], 1):
        c = ws_m.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.border = thin_border

    dv_status_m = DataValidation(type="list", formula1=f'"{status_options}"', allow_blank=True)
    ws_m.add_data_validation(dv_status_m)

    for i, m in enumerate(result.milestones, 2):
        ws_m.cell(row=i, column=1, value=m.get("week", ""))
        ws_m.cell(row=i, column=2, value=m.get("name", ""))
        ws_m.cell(row=i, column=3, value=m.get("description", ""))
        ws_m.cell(row=i, column=4, value="Pendiente")
        for col in range(1, 5):
            ws_m.cell(row=i, column=col).border = thin_border

    dv_status_m.add(f"D2:D{len(result.milestones) + 1}")
    for col_letter, w in zip("ABCD", (10, 30, 60, 14)):
        ws_m.column_dimensions[col_letter].width = w

    # ════════════════════════════════════════════════════════════
    # HOJA 6: ESTRUCTURA
    # ═════════════════════════════════════════════════════��══════
    ws_s = wb.create_sheet("Estructura")
    ws_s["A1"] = "Estructura documental sugerida"
    ws_s["A1"].font = Font(bold=True, size=14, color="1F4E78")
    ws_s["A3"] = result.suggested_structure
    ws_s["A3"].alignment = Alignment(wrap_text=True, vertical="top")
    ws_s.column_dimensions["A"].width = 100
    ws_s.row_dimensions[3].height = 400

    # ════════════════════════════════════════════════════════════
    # HOJA 7: INSTRUCCIONES
    # ════════════════════════════════════════════════════════════
    ws_help = wb.create_sheet("📖 Instrucciones")
    instructions = [
        ("CÓMO USAR ESTE EXCEL", ""),
        ("", ""),
        ("HOJA 'Seguimiento'", "Es tu herramienta principal de gestión. Actualiza el ESTADO y % AVANCE de cada tarea."),
        ("  → Estado", "Usa el desplegable: Pendiente, En Plan, En Progreso, Completado, Bloqueado, Retrasado, En Revisión, Cancelado"),
        ("  → % Avance", "Introduce un número entre 0 y 100. Los colores cambian automáticamente."),
        ("  → Dependencias", "Escribe los IDs de tareas predecesoras separados por coma (ej: T-001, T-003)"),
        ("  → Notas", "Añade cualquier comentario o bloqueante"),
        ("", ""),
        ("HOJA 'Gantt Chart'", "Visualización temporal de las tareas. Las barras azules muestran la duración planificada."),
        ("  → Para modificar duraciones", "Cambia 'Semana inicio' y 'Semana fin' en la hoja Seguimiento. El Gantt NO se actualiza solo (es estático)."),
        ("  → Si necesitas regenerar el Gantt", "Modifica los valores en Seguimiento y vuelve a generar desde la herramienta."),
        ("  → Hitos", "Aparecen como diamantes naranjas (◆) en su semana correspondiente."),
        ("  → Estado sincronizado", "La columna Estado del Gantt se lee automáticamente de la hoja Seguimiento (fórmula)."),
        ("", ""),
        ("HOJA 'Entregables'", "Lista de entregables con estado y responsable editables."),
        ("", ""),
        ("COLORES DE ESTADO", ""),
        ("  Pendiente", "Azul claro — aún no arrancada"),
        ("  En Plan", "Gris azulado — planificada pero no iniciada"),
        ("  En Progreso", "Verde claro — en ejecución"),
        ("  Completado", "Verde fuerte — finalizada"),
        ("  Bloqueado", "Rojo claro — necesita desbloquearse"),
        ("  Retrasado", "Naranja — fuera de plazo"),
        ("  En Revisión", "Amarillo — pendiente de aprobación"),
        ("  Cancelado", "Gris — descartada"),
    ]
    for i, (col1, col2) in enumerate(instructions, 1):
        ws_help.cell(row=i, column=1, value=col1)
        ws_help.cell(row=i, column=2, value=col2)
        if col1 and not col1.startswith(" "):
            ws_help.cell(row=i, column=1).font = Font(bold=True, size=12 if i == 1 else 11)
    ws_help.column_dimensions["A"].width = 32
    ws_help.column_dimensions["B"].width = 90

    wb.save(out_path)
    return out_path


# ─────────────────────────────────────────────────────────────
# 3. ONBOARDING KIT
# ─────────────────────────────────────────────────────────────

@dataclass
class OnboardingResult:
    person_role: str
    project: str
    welcome_doc_md: str
    glossary: list[dict]  # [{term, definition}]
    recent_decisions: list[str]
    first_tasks: list[dict]  # [{task, week, description}]


def generate_onboarding_kit(
    project: str,
    person_role: str,
    person_name: str = "",
    start_date: str = "",
) -> OnboardingResult:
    """
    Genera un kit de onboarding personalizado para un nuevo integrante.
    """
    keywords = _extract_keywords(f"{project} {person_role} {person_name}")
    project_context = _gather_project_context(
        project,
        max_actas=5,
        relevance_keywords=keywords,
        budget_chars=BUDGET["onboarding_project"],
    )
    global_specs = _gather_global_specs_tiered(
        relevance_keywords=keywords,
        exclude_project=project,
        budget_chars=BUDGET["onboarding_global"],
    )

    prompt = f"""You are a senior PMO creating an onboarding kit for a new team member. Write everything in Spanish.

## NEW TEAM MEMBER
- Name: {person_name or 'Nuevo integrante'}
- Role: {person_role}
- Start date: {start_date or 'Next week'}
- Project: {project}

## PROJECT CONTEXT — Tier 1 (actas recientes y specs del proyecto):
{project_context}

## CROSS-PROJECT SPECS — Tier 2 (resúmenes de otros proyectos del equipo):
{global_specs}

## TASK
Generate a comprehensive onboarding kit including:
1. A welcome document (markdown) with project overview, team structure hints, and key contacts
2. A glossary of technical terms and acronyms used in this project
3. Summary of the last 3 key decisions that affect this person's role
4. Assigned tasks for the first 15 days (progressive difficulty)

## OUTPUT FORMAT (JSON only):
{{
  "welcome_doc_md": "# Bienvenida al proyecto {{project}}\\n\\n...(full markdown document)",
  "glossary": [
    {{"term": "KDD", "definition": "Knowledge-Driven Development..."}},
    {{"term": "ADR", "definition": "Architecture Decision Record..."}}
  ],
  "recent_decisions": [
    "Se decidió usar Vertex AI para...",
    "Se aprobó la arquitectura de..."
  ],
  "first_tasks": [
    {{"task": "Leer specs del proyecto", "week": 1, "description": "..."}},
    {{"task": "Configurar entorno", "week": 1, "description": "..."}}
  ]
}}
"""
    raw = call_github_models(prompt)
    cleaned = raw.strip()
    if cleaned.startswith("```"):
        cleaned = re.sub(r"^```(?:json)?\s*\n?", "", cleaned)
        cleaned = re.sub(r"\n?```\s*$", "", cleaned)

    data = json.loads(cleaned)
    return OnboardingResult(
        person_role=person_role,
        project=project,
        welcome_doc_md=data.get("welcome_doc_md", ""),
        glossary=data.get("glossary", []),
        recent_decisions=data.get("recent_decisions", []),
        first_tasks=data.get("first_tasks", []),
    )


def onboarding_to_txt(result: OnboardingResult, person_name: str = "") -> str:
    """Serializa un OnboardingResult a texto plano para descarga."""
    lines = []
    lines.append(f"ONBOARDING KIT — Proyecto {result.project}")
    lines.append(f"Generado: {date.today().isoformat()}")
    if person_name:
        lines.append(f"Para: {person_name}")
    lines.append(f"Rol: {result.person_role}")
    lines.append("=" * 70)
    lines.append("")
    lines.append("DOCUMENTO DE BIENVENIDA")
    lines.append("-" * 70)
    lines.append(result.welcome_doc_md)
    lines.append("")
    lines.append("=" * 70)
    lines.append("GLOSARIO DE TÉRMINOS")
    lines.append("-" * 70)
    for item in result.glossary:
        lines.append(f"  • {item.get('term', '')}: {item.get('definition', '')}")
    lines.append("")
    lines.append("=" * 70)
    lines.append("DECISIONES CLAVE RECIENTES")
    lines.append("-" * 70)
    for i, d in enumerate(result.recent_decisions, 1):
        lines.append(f"  {i}. {d}")
    lines.append("")
    lines.append("=" * 70)
    lines.append("TAREAS PARA LOS PRIMEROS 15 DÍAS")
    lines.append("-" * 70)
    for t in result.first_tasks:
        lines.append(f"  [Semana {t.get('week', '?')}] {t.get('task', '')}")
        if t.get("description"):
            lines.append(f"      → {t.get('description')}")
    return "\n".join(lines)







